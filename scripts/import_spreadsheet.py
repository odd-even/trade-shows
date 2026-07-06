#!/usr/bin/env python3
"""Import Trade Show Master 2026 spreadsheet into data/shows.json.

Parses the JF show-tab template:
  Col A: task/label | B: info | C: notes | D: WHO'S RESPONSIBLE | E: WHEN'S IT DUE | F: IS IT DONE?

Usage:
  python scripts/import_spreadsheet.py
  python scripts/import_spreadsheet.py path/to/Trade\ Show\ Master\ 2026.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "source" / "Trade Show Master 2026.xlsx"
FALLBACK_XLSX = ROOT / "data" / "Trade Show Master 2026.xlsx"
OUTPUT = ROOT / "data" / "shows.json"

SKIP_SHEETS = {
    "summary",
    "summary compilation",
    "assumptions",
    "standard equipment",
    "copy me template",
    "master template",
    "cast",
    "nb hort",
    "tnla",
    "igc east",
    "sobeys",
    "grn trd sk",
    "negc",
    "ngc",
    "great lakes",
    "tpie",
}

SECTION_MARKERS = {
    "SHOW PURCHASE",
    "PERSONNEL",
    "BOOTH SETUP",
    "FREIGHT",
    "LITERATURE",
    "ON SITE",
    "POST SHOW REVIEW",
}

# Map spreadsheet owner → current team (per June 2026 handoff email)
OWNER_CURRENT = {
    "artroom": "Graphics",
    "elisabeth": None,  # resolved by section
    "elisabeth/reps": "Debbie",
    "mailroom": "Mailroom",
    "michael j": "Michael",
    "n/a": None,
    "personnel": "On-site reps",
    "peter": "Peter",
}

SECTION_OWNER_OVERRIDE = {
    "PERSONNEL": "Debbie",
    "SHOW PURCHASE": "Graphics",
    "BOOTH SETUP": "Graphics",
    "FREIGHT": "Graphics",
    "LITERATURE": "Graphics",
    "ON SITE": "Graphics",
    "POST SHOW REVIEW": "Peter",
}

BOOTH_SETUP_OVERRIDES = {
    "design": "Graphics",
    "design notes": "Graphics",
    "pack display": "Graphics",
    "book plugs": "Peter",
    "book finished": "Peter",
}

FREIGHT_MICHAEL_TASKS = {
    "carrier",
    "freight departure",
    "end of show pick up",
    "transportation",
    "show site arrival",
}


def slugify(name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-")
    return slug or "show"


def parse_date(value: object) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%b %d, %Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def is_done(value: object) -> bool:
    if value is None:
        return False
    raw = str(value).strip().lower()
    return raw in {"yes", "y", "done", "complete", "completed", "x", "paid"}


def normalize_owner(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text.lower() in {"n/a", "na", ""}:
        return None
    lower = text.lower()
    if lower.startswith("michael"):
        return "Michael"
    if lower.startswith("elisabeth"):
        return "Elisabeth/Reps" if "/" in lower else "Elisabeth"
    return text


def resolve_current_owner(task_name: str, section: str, spreadsheet_owner: str | None) -> str | None:
    if not spreadsheet_owner:
        return SECTION_OWNER_OVERRIDE.get(section)

    key = spreadsheet_owner.strip().lower()
    task_lower = task_name.lower()

    if key == "elisabeth":
        if section == "PERSONNEL":
            return "Debbie"
        if section == "BOOTH SETUP":
            for prefix, owner in BOOTH_SETUP_OVERRIDES.items():
                if task_lower.startswith(prefix):
                    return owner
            return "Graphics"
        if section == "FREIGHT":
            for prefix in FREIGHT_MICHAEL_TASKS:
                if prefix in task_lower:
                    return "Michael"
            return "Graphics"
        return SECTION_OWNER_OVERRIDE.get(section, "Graphics")

    if key in OWNER_CURRENT:
        mapped = OWNER_CURRENT[key]
        if mapped:
            return mapped
    return spreadsheet_owner


def parse_status(done_val: object, notes: str | None) -> str:
    if is_done(done_val):
        return "completed"
    if notes and "not attending" in notes.lower():
        return "cancelled"
    if notes and "out of business" in notes.lower():
        return "cancelled"
    return "pending"


def parse_show_sheet(ws, sheet_name: str) -> dict | None:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 10:
        return None

    show_name = sheet_name.strip()
    for row in rows[:3]:
        if row and row[0] == "Show" and row[1]:
            show_name = str(row[1]).strip()

    facts: dict[str, object] = {}
    tasks: list[dict] = []
    section = ""
    show_id = slugify(show_name)

  # status from row 0 col 2-3 (e.g. "Out of business")
    header_note = rows[0][2] if rows[0] and len(rows[0]) > 2 else None
    show_status = "upcoming"
    if header_note:
        note_l = str(header_note).lower()
        if "out of business" in note_l or "not attending" in note_l:
            show_status = "cancelled"
        elif "canceled" in note_l or "cancelled" in note_l:
            show_status = "cancelled"

    for row_idx, row in enumerate(rows, start=1):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        label_upper = label.upper()

        if label_upper in SECTION_MARKERS:
            section = label_upper
            continue

        if label_upper == "GENERAL FACTS":
            section = "GENERAL"
            continue

        if section == "GENERAL" and label in {
            "Show Start Date",
            "Show End Date",
            "Where",
            "Booth Size",
            "Booth Number(s)",
            "Booth Numbers",
            "Exhibitor Resource Center",
            "Product Focus",
        }:
            key = {
                "Show Start Date": "start",
                "Show End Date": "end",
                "Where": "location",
                "Booth Size": "boothSize",
                "Booth Number(s)": "booth",
                "Booth Numbers": "booth",
                "Exhibitor Resource Center": "exhibitorPortal",
                "Product Focus": "productFocus",
            }[label]
            val = row[1]
            if key in {"start", "end"}:
                facts[key] = parse_date(val)
            elif val is not None:
                facts[key] = str(val).strip() if not isinstance(val, (int, float)) else val
            continue

        if section in SECTION_MARKERS and label_upper not in SECTION_MARKERS and label != "INFO":
            info = str(row[1]).strip() if row[1] is not None else None
            notes = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            owner_raw = normalize_owner(row[3] if len(row) > 3 else None)
            due = parse_date(row[4] if len(row) > 4 else None)
            done_val = row[5] if len(row) > 5 else None

            if not info and not notes and not owner_raw and not due and not done_val:
                continue

            owner_current = resolve_current_owner(label, section, owner_raw)
            status = parse_status(done_val, notes)

            task = {
                "id": f"{show_id}-{row_idx}",
                "task": label,
                "section": section.replace(" ", "_").lower(),
                "ownerSpreadsheet": owner_raw,
                "owner": owner_current,
                "status": status,
                "source": f"spreadsheet:{sheet_name.strip()}",
            }
            if info:
                task["info"] = info
            if notes:
                task["notes"] = notes
            if due:
                task["dueDate"] = due

            label_lower = label.lower()
            if not task.get("dueDate") and info:
                if "advance shipping opens" in label_lower or "advance shipping closes" in label_lower:
                    info_date = parse_date(info)
                    if info_date:
                        task["dueDate"] = info_date

            if section == "SHOW PURCHASE" and label == "Liability Insurance":
                task["priority"] = "critical"
            elif section == "FREIGHT" and "departure" in label_lower:
                task["priority"] = "critical"
            elif section == "FREIGHT" and "advance shipping closes" in label_lower:
                task["priority"] = "critical"
            elif status == "pending" and section in {"SHOW PURCHASE", "PERSONNEL", "BOOTH SETUP", "FREIGHT"}:
                task["priority"] = "high"

            tasks.append(task)

    if not facts.get("start") and not tasks:
        return None

    dates = {"start": facts.get("start"), "end": facts.get("end")}

    show = {
        "id": show_id,
        "name": show_name,
        "year": 2026,
        "dates": dates,
        "location": facts.get("location"),
        "boothSize": facts.get("boothSize"),
        "booth": facts.get("booth"),
        "exhibitorPortal": facts.get("exhibitorPortal"),
        "productFocus": facts.get("productFocus"),
        "status": show_status,
        "tasks": tasks,
    }
    return show


def import_workbook(path: Path) -> dict:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("Install openpyxl: python3 -m pip install openpyxl") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    shows: list[dict] = []
    imported: list[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        show = parse_show_sheet(ws, sheet_name)
        if show and (show.get("tasks") or show.get("dates", {}).get("start")):
            shows.append(show)
            imported.append(sheet_name.strip())

    shows.sort(key=lambda s: (s.get("dates") or {}).get("start") or "9999")

    return {
        "lastUpdated": date.today().isoformat(),
        "year": 2026,
        "source": f"Trade Show Master 2026.xlsx ({path.name})",
        "teams": {
            "Graphics": "Booth registration/payment, COI, electric, accessories, freight coordination, artroom design & booth graphics",
            "Debbie": "Personnel registration, hotels, flights, transport, petty cash & credit cards (Karen), travel insurance",
            "Michael": "Truck shipping — departure, arrival, pickup",
            "Peter": "Sales staff assignments, plant booking, post-show review",
            "Mailroom": "Contact books, pack literature",
        },
        "ownerLegend": {
            "ownerSpreadsheet": "Original owner from spreadsheet (legacy Elisabeth assignments)",
            "owner": "Current responsible person/team after June 2026 handoff",
        },
        "shows": shows,
        "importedTabs": imported,
    }


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        xlsx = FALLBACK_XLSX
    if not xlsx.exists():
        print(f"Spreadsheet not found. Place file at:\n  {DEFAULT_XLSX}")
        sys.exit(1)

    data = import_workbook(xlsx)
    OUTPUT.write_text(json.dumps(data, indent=2) + "\n")

    pending = sum(1 for s in data["shows"] for t in s["tasks"] if t["status"] == "pending")
    print(f"Imported {len(data['importedTabs'])} shows, {pending} open tasks → {OUTPUT}")
    for name in data["importedTabs"]:
        print(f"  • {name}")


if __name__ == "__main__":
    main()
